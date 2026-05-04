"""
Formatar Excel do Relatório TBDC - Identidade Visual Latitude Genética
Cores: #00BB31 (verde) e #464646 (cinza escuro)

Uso: python formatar_excel.py <caminho_do_arquivo.xlsx>
Ou arraste o arquivo .xlsx em cima deste script.

Formatação aplicada:
- Header: fundo verde #00BB31, texto branco, negrito, centralizado
- Dados: fonte Calibri 10pt, cor #464646
- Zebra striping: linhas alternadas com verde claro #F0FFF0
- Colunas auto-dimensionadas
- Auto-filtro no header
- Header congelado (freeze pane)
- Bordas finas cinza
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def formatar_relatorio(filepath):
    print(f"[*] Abrindo: {filepath}")
    wb = load_workbook(filepath)
    ws = wb.active

    # === CORES LATITUDE ===
    VERDE = "00BB31"
    CINZA = "464646"
    VERDE_CLARO = "F0FFF0"
    BRANCO = "FFFFFF"

    # === ESTILOS ===
    header_font = Font(name="Calibri", size=11, bold=True, color=BRANCO)
    header_fill = PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(bottom=Side(style="thin", color=CINZA))

    data_font = Font(name="Calibri", size=10, color=CINZA)
    data_align = Alignment(vertical="center")
    data_border = Border(bottom=Side(style="hair", color="D9D9D9"))

    zebra_fill = PatternFill(start_color=VERDE_CLARO, end_color=VERDE_CLARO, fill_type="solid")

    # === LARGURAS OTIMIZADAS ===
    col_widths = {
        'A': 10, 'B': 14, 'C': 16, 'D': 14, 'E': 35,    # ID Campo → Nome Fazenda
        'F': 40, 'G': 18, 'H': 10, 'I': 8, 'J': 28,      # Proprietário → Rep Técnico
        'K': 30, 'L': 18, 'M': 10, 'N': 16,               # Protocolo → ID Tratamento
        'O': 12, 'P': 30, 'Q': 18, 'R': 10,               # ID Cultivar → GM
        'S': 14, 'T': 16, 'U': 25, 'V': 16,               # Área → Estado
        'W': 16, 'X': 25, 'Y': 28,                        # Tipo cultivar → Genética
        'Z': 18, 'AA': 10, 'AB': 14, 'AC': 18             # Altura → ID Campo Avaliado
    }

    print(f"[+] Processando {ws.max_row} linhas x {ws.max_column} colunas...")

    # === APLICAR LARGURAS ===
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # === FORMATAR HEADER (linha 1) ===
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # === FORMATAR DADOS ===
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_align
            cell.border = data_border
            # Zebra striping
            if row_idx % 2 == 0:
                cell.fill = zebra_fill

    # === AUTO-FILTRO ===
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col}1"

    # === CONGELAR HEADER ===
    ws.freeze_panes = "A2"

    # === ALTURA DAS LINHAS ===
    ws.row_dimensions[1].height = 30  # Header mais alto

    # === SALVAR ===
    output_path = filepath  # Sobrescreve o original
    wb.save(output_path)
    wb.close()

    print(f"[OK] Formatado com sucesso: {output_path}")
    print(f"     Header verde #00BB31 + texto branco")
    print(f"     {ws.max_column} colunas dimensionadas")
    print(f"     Auto-filtro + header congelado")
    print(f"     Zebra striping nas linhas de dados")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        # Procura o arquivo mais recente na pasta padrão
        pasta = r"C:\Users\leonardo.alves\Claude\Relatorio Avaliacoes"
        arquivos = [
            os.path.join(pasta, f) for f in os.listdir(pasta)
            if f.startswith("relatorio-avaliacoes-") and f.endswith(".xlsx")
        ]
        if arquivos:
            filepath = max(arquivos, key=os.path.getmtime)
            print(f"[?] Nenhum arquivo informado. Usando mais recente: {os.path.basename(filepath)}")
        else:
            print("[!] Uso: python formatar_excel.py <arquivo.xlsx>")
            sys.exit(1)
    else:
        filepath = sys.argv[1]

    if not os.path.exists(filepath):
        print(f"[!] Arquivo nao encontrado: {filepath}")
        sys.exit(1)

    formatar_relatorio(filepath)
